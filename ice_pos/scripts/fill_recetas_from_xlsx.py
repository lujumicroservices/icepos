#!/usr/bin/env python3
"""
Lee Recetas RN 2025.xlsx y actualiza recetas_formato.json con ingredientes
mapeados a productos existentes. Requiere: pip install openpyxl (en venv).
Uso: python3 scripts/fill_recetas_from_xlsx.py [ruta_xlsx] [ruta_json]
"""
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instala openpyxl: pip install openpyxl")
    sys.exit(1)

# Mapeo nombre en Excel -> nombre exacto en recetas_formato.json
PRODUCT_ALIASES = {
    "jamón pechuga de pavo": "Jamón de Pechuga de Pavo",
    "jamon pechuga de pavo": "Jamón de Pechuga de Pavo",
    "baguette jamon de pechuga de pavo": "Jamón de Pechuga de Pavo",
    "baguette jamón pechuga pavo": "Jamón de Pechuga de Pavo",
    "pizza": "Pizza",
    "vegetariano": "Vegetariano",
    "light": "Light",
    "español": "Español",
    "baguette español": "Español",
    "lomo canadiense": "Lomo Canadiense",
    "baguette de lomo canadiense": "Lomo Canadiense",
    "alemán": "Alemán",
    "baguette aleman": "Alemán",
    "italiano": "Italiano",
    "baguette italiano": "Italiano",
    "mango & ginger power": "Mango & Ginger Power",
    "mango and ginger power": "Mango & Ginger Power",
    "shake matcha": "Shake Matcha Mediano",
    "shake taro": "Shake Taro Mediano",
    "shakes taro / carbon": "Shake Taro Mediano",
    "pink flamingo": "Shake Pink Flamingo Mediano",
    "soda organica": "Soda Orgánica Regular Mediano",
    "soda organica ": "Soda Orgánica Regular Mediano",
    "soda organica regular": "Soda Orgánica Regular Mediano",
    "soda organica+ nieve de vainilla": "Soda Orgánica con Nieve Mediano",
    "soda organica con nieve": "Soda Orgánica con Nieve Mediano",
    "sandwich rey": "Sándwich Rey",
    "sándwich rey": "Sándwich Rey",
    "reyes waffle": "Reyes Waffle",
    "bubble waffle": "Reyes Waffle",
    "paletas para perritos": "Paletas para Perritos Chica",
    "paletas para perritos chica": "Paletas para Perritos Chica",
    "paletas para perritos grande": "Paletas para Perritos Grande",
    "frappe pan de muerto": "Frapuccino Mediano",
    "frappe pay calabaza": "Frapuccino Mediano",
    "cappuccino": "Capuccino Mediano",
    "café americano": "Americano Mediano",
    "americano": "Americano Mediano",
    "mateadas": "Malteada Mediano",
    "malteada": "Malteada Mediano",
    "cold chai": "Cold Chai Mediano",
    "frappuccino": "Frapuccino Mediano",
    "late frio": "Latte Frío Mediano",
    "croissant de jamon y queso": "croissant de jamon y queso",
    "croissant de nutella": "croissant de nutella",
}

# Normalizar insumo Excel -> insumo en insumos_guia / recetas
INSUMO_ALIASES = {
    "jarabe": "Jarabe",
    "jarabe (caramel/vainilla)": "Jarabe",
    "leche": "Leche",
    "café": "Café",
    "cafe": "Café",
    "agua": "Agua mineral",
    "agua mineral": "Agua mineral",
    "agua fria": "Agua mineral",
    "hielo": "Hielo",
    "cubos de hielo": "Hielo",
    "concentrado soda": "Concentrado soda",
    "concentrado a vaso": "Concentrado soda",
    "vaso shake": "Vaso Mediano",
    "vaso malteada": "Vaso Malteada",
    "vaso mediano": "Vaso Mediano",
    "vaso grande": "Vaso Grande",
    "base taro": "Base Taro",
    "base taro o carbon": "Base Taro",
    "base pan de muerto": "Base Pan de Muerto",
    "paq. mango congelado": "Mango congelado",
    "mango congelado (80gr)": "Mango congelado",
    "gengibre polvo": "Jengibre polvo",
    "medidas gengibre polvo": "Jengibre polvo",
    "leche de almendra": "Leche de almendras",
    "pan": "Pan baguette",
    "pan cortar por mitad": "Pan baguette",
    "queso cheddar": "Queso cheddar",
    "queso cheddar cortar en 2": "Queso cheddar",
    "jamon pavo": "Pechuga de pavo",
    "jamon pavo 2 rebanadas": "Pechuga de pavo",
    "mayonesa": "Mayonesa",
    "mayonesa untar": "Mayonesa",
    "pepperoni": "Pepperoni",
    "jamón selva negra": "Jamón selva negra",
    "jamón serrano": "Jamón selva negra",
    "queso gouda": "Queso gouda",
    "jitomate": "Jitomate",
    "cebolla": "Cebolla",
    "cebolla morada": "Cebolla",
    "lechuga": "Lechuga",
    "salsa de frambuesa": "Aderezo frambuesa",
    "aderezo de cilantro": "Cilantro",
    "cilantro": "Cilantro",
    "pepino": "Pepino",
    "rodajas de pepino": "Pepino",
    "panela": "Queso panela",
    "queso panela": "Queso panela",
    "servilleta": "Servilleta",
    "popote": "Popote",
    "tapa": "Tapa domo",
    "tapa domo": "Tapa domo",
    "tapa lisa": "Tapa lisa",
    "crema batida": "Crema batida",
    "muffin": "Muffin",
    "polvo base malteada": "Polvo base malteada",
    "baguetera": "Baguetera",
    "papel encerado": "Papel encerado",
    "aluminio": "Aluminio",
    "platanitos": "Platanitos",
    "jalapeños": "Jalapeños",
    "kiwis": "Kiwi",
    "pan cortar por mitad": "Pan baguette",
    "queso cheddar cortar en 2": "Queso cheddar",
    "jamon pavo 2 rebanadas": "Pechuga de pavo",
    "jamon pavo": "Pechuga de pavo",
    "mayonesa untar": "Mayonesa",
    "mayonesa untar sobre proteina": "Mayonesa",
    "papel aluminio": "Aluminio",
    "aderezo cilantro": "Cilantro",
    "queso gouda cortar en 2": "Queso gouda",
    "peperoni 5 rodajas": "Pepperoni",
    "aderezo chipotle": "Salsa chipotle",
    "kiwi": "Kiwi",
    "azucar": "Azúcar",
    "azúcar": "Azúcar",
    "galletas": "Galletas",
    "nieve": "Nieve",
    "waffle": "Waffle",
    "canasta cartón": "Canasta cartón",
}


def normalize(s):
    if s is None or (isinstance(s, float) and (s != s or s == 0)):
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def parse_quantity_and_name(text):
    """Extrae cantidad y nombre de insumo de una línea tipo '4 Medida Jarabe', '1 LT DE AGUA', '60 GR DE AZUCAR', '1½ Base Taro', 'Jamon Pavo 2 rebanadas'."""
    if not text or not str(text).strip():
        return None
    text = str(text).strip()
    # "X rebanadas" / "X rodajas" al final
    m = re.match(r"^(.+?)\s+(\d+)\s*(?:rebanadas?|rodajas?)$", text, re.I)
    if m:
        return (normalize_insumo(m.group(1)), float(m.group(2)))
    # Número al inicio (incl. 1½, 1.5, etc.)
    m = re.match(r"^(\d+[\.,]?\d*)\s*(?:medidas?\s+de\s+)?(?:de\s+)?(.+)$", text, re.I)
    if m:
        qty_str, name = m.group(1).replace(",", "."), m.group(2).strip()
        try:
            if "½" in qty_str or "1/2" in qty_str:
                qty = 0.5 if "1" not in qty_str or qty_str.startswith("½") else 1.5
            else:
                qty = float(qty_str)
        except ValueError:
            qty = 1.0
        return (normalize_insumo(name), qty)
    m = re.match(r"^(\d+)\s*(?:gr|g|gramos?)\s+(?:de\s+)?(.+)$", text, re.I)
    if m:
        return (normalize_insumo(m.group(2)), float(m.group(1)) / 1000.0)
    m = re.match(r"^(\d+)\s*(?:lt|litros?|l)\s+(?:de\s+)?(.+)$", text, re.I)
    if m:
        return (normalize_insumo(m.group(2)), float(m.group(1)))
    m = re.match(r"^(\d+)\s*(.+)$", text, re.I)
    if m:
        try:
            return (normalize_insumo(m.group(2)), float(m.group(1)))
        except ValueError:
            pass
    return (normalize_insumo(text), 1.0)


def normalize_insumo(name):
    n = normalize(name)
    for k, v in INSUMO_ALIASES.items():
        if k in n or n in k:
            return v
    return name.strip().title() if name else ""


def extract_recipes_from_sheets(wb):
    """Extrae por hoja: lista de (producto_normalized, ingredientes_list)."""
    recipes_by_product = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if sheet_name == "P. Bebidas Calientes":
            current_product = None
            for row in rows[2:]:
                name_cell = (row[0] or "").strip() if row else ""
                prep_cell = (row[2] or "").strip() if len(row) > 2 else ""
                if name_cell and not name_cell.upper().startswith("NOMBRE"):
                    if "(" not in name_cell and name_cell not in ("Mediano", "Grande"):
                        current_product = name_cell
                        recipes_by_product.setdefault(normalize(current_product), []).clear()
                    if current_product and prep_cell and not prep_cell.startswith("Seleccionar"):
                        ing = parse_quantity_and_name(prep_cell)
                        if ing and ing[0]:
                            recipes_by_product.setdefault(normalize(current_product), []).append(ing)
        elif sheet_name == "P. Babidas Frias" or sheet_name == "P. Fit Shakes":
            for i, row in enumerate(rows):
                name_cell = (row[0] or "").strip() if row else ""
                if name_cell and "Shake" in name_cell or "Mango" in name_cell or "Pink" in name_cell or "Taro" in name_cell or "Matcha" in name_cell or "Carbon" in name_cell:
                    product_key = normalize(name_cell.split("/")[0].strip())
                    prep_cols = [row[j] for j in range(2, min(8, len(row))) if row[j]]
                    for cell in prep_cols:
                        if cell and str(cell).strip() and not isinstance(cell, (int, float)):
                            ing = parse_quantity_and_name(str(cell))
                            if ing and ing[0]:
                                recipes_by_product.setdefault(product_key, []).append(ing)
        elif sheet_name == "Soda Organica":
            current = None
            for row in rows:
                name_cell = (row[0] or "").strip() if row else ""
                prep = (row[2] or row[3] or "").strip() if len(row) > 3 else ""
                if "Soda" in (name_cell or ""):
                    current = normalize(name_cell)
                if current and prep and "Llenar" in prep:
                    recipes_by_product.setdefault(current, []).append(("Concentrado soda", 0.07))
                if current and prep and ("hielo" in prep.lower() or "cubos" in prep.lower()):
                    recipes_by_product.setdefault(current, []).append(("Hielo", 1))
                if current and prep and "agua" in prep.lower():
                    recipes_by_product.setdefault(current, []).append(("Agua mineral", 0.35))
        elif sheet_name == "P. Bagettes":
            current_key = None
            for row in rows[3:]:
                name_cell = (row[0] or "").strip() if row else ""
                prep = (row[1] or "").strip() if len(row) > 1 else ""
                if name_cell and ("Baguette" in name_cell or "baguette" in name_cell.lower()):
                    key = normalize(name_cell)
                    if "pechuga" in key or "pavo" in key or "jamon" in key:
                        current_key = "jamón de pechuga de pavo"
                    elif "pizza" in key:
                        current_key = "pizza"
                    else:
                        current_key = key.replace("baguette", "").strip()
                    recipes_by_product.setdefault(current_key, []).clear()
                elif current_key and prep and not prep.upper().startswith("METER") and "Grill" not in prep:
                    ing = parse_quantity_and_name(prep)
                    if ing and ing[0]:
                        recipes_by_product.setdefault(current_key, []).append(ing)
        elif sheet_name == "Baguettes Gourtmet":
            for i in range(2, min(len(rows), 80), 4):
                row = rows[i] if i < len(rows) else []
                name_cell = (row[1] if len(row) > 1 else "") or ""
                if isinstance(name_cell, str) and "Baguette" in name_cell:
                    short = name_cell.replace("Baguette", "").replace("de ", "").strip()
                    key = normalize(short)
                    if key and key not in ("gourtmet", "gourmet"):
                        recipes_by_product.setdefault(key, []).extend([
                            ("Pan baguette", 1), ("Baguetera", 1), ("Papel encerado", 1), ("Aluminio", 1),
                            ("Servilleta", 1), ("Platanitos", 1), ("Jalapeños", 1),
                        ])
        elif sheet_name == "P. AGUAS FRESCAS":
            for i, row in enumerate(rows):
                name_cell = (row[0] or "").strip() if row else ""
                if name_cell and "KIWI" in name_cell.upper() and "PARA" not in name_cell.upper():
                    product_key = "agua kiwi"
                    for j in range(i + 1, min(i + 10, len(rows))):
                        r = rows[j]
                        cell = (r[0] or "").strip() if r else ""
                        if cell and re.match(r"^\d+\s*(KIWIS?|LT|GR|GR\.)", cell, re.I):
                            ing = parse_quantity_and_name(cell)
                            if ing and ing[0]:
                                recipes_by_product.setdefault(product_key, []).append(ing)
                    break
        elif sheet_name == "PARA EL ANTOJO":
            for row in rows[2:]:
                name_cell = (row[0] or "").strip() if row else ""
                if not name_cell:
                    continue
                n = normalize(name_cell)
                if "sandwich rey" in n or "sándwich rey" in n:
                    recipes_by_product.setdefault("sándwich rey", []).extend([
                        ("Galletas", 2), ("Nieve", 0.05), ("Canasta cartón", 1), ("Papel encerado", 1), ("Servilleta", 1),
                    ])
                elif "reyes waffle" in n or "waffle" in n and "bubble" not in n:
                    recipes_by_product.setdefault("reyes waffle", []).extend([
                        ("Waffle", 2), ("Nieve", 0.1), ("Crema batida", 1), ("Baguetera", 1), ("Servilleta", 1),
                    ])
                elif "paletas" in n and "perrit" in n:
                    if "chica" in n or "CHICA" in name_cell:
                        recipes_by_product.setdefault("paletas para perritos chica", []).append(("Paleta Agua", 1))
                    elif "grande" in n or "GRANDE" in name_cell:
                        recipes_by_product.setdefault("paletas para perritos grande", []).append(("Paleta Agua", 1))
    return recipes_by_product


def main():
    base = Path(__file__).resolve().parent.parent
    path_xlsx = sys.argv[1] if len(sys.argv) > 1 else str(Path.home() / "Downloads/Recetas RN 2025.xlsx")
    path_json = sys.argv[2] if len(sys.argv) > 2 else str(base / "assets/data/recetas_formato.json")
    if not Path(path_xlsx).exists():
        print("No se encontró:", path_xlsx)
        sys.exit(1)
    wb = openpyxl.load_workbook(path_xlsx, read_only=True, data_only=True)
    recipes_from_xlsx = extract_recipes_from_sheets(wb)
    wb.close()
    with open(path_json, "r", encoding="utf-8") as f:
        data = json.load(f)
    recetas = data.get("recetas", [])
    updated = 0
    for r in recetas:
        prod = r.get("producto", "")
        if not prod:
            continue
        current_ing = r.get("ingredientes") or []
        if current_ing:
            continue
        key = normalize(prod)
        matched_ingredients = None
        for alias, json_name in PRODUCT_ALIASES.items():
            if json_name == prod:
                matched_ingredients = recipes_from_xlsx.get(alias) or recipes_from_xlsx.get(key)
                break
        if not matched_ingredients:
            matched_ingredients = recipes_from_xlsx.get(key)
        if matched_ingredients:
            seen = set()
            unique = []
            for ins, qty in matched_ingredients:
                if ins and ins not in seen:
                    seen.add(ins)
                    unique.append({"insumo": ins, "cantidad": round(qty, 4) if isinstance(qty, float) else qty})
            if unique:
                r["ingredientes"] = unique
                updated += 1
                print("Actualizado:", prod, "->", len(unique), "ingredientes")
    with open(path_json, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print("Total recetas actualizadas:", updated)


if __name__ == "__main__":
    main()
